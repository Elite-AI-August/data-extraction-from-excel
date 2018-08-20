import openpyxl
import csv
import datetime
import win32com.client

def refresh_cc_sheets(path,input_file,countries):
# Open Excel
    
    Application = win32com.client.Dispatch("Excel.Application")
 
 # Show Excel. While this is not required, it can help with debugging
    Application.Visible = 1
    Application.DisplayAlerts=False
    Application.AskToUpdateLinks = False

    for  country in countries.values():
 # Open Your Workbook
        Workbook = Application.Workbooks.open(path + input_file + country + '.xlsx')
        try:
            Workbook.UpdateLink(Name=Workbook.LinkSources())

        except Exception as e:
            print(e)
        # Refesh All
        Workbook.RefreshAll()
        Application.CalculateUntilAsyncQueriesDone()
        print(country + ' - Done')
    # Saves the Workbook
        Workbook.Save()
        Workbook.Close()
    Application.Visible = 0
    Application.DisplayAlerts=True
    Application.AskToUpdateLinks = True
 # Closes Excel
    Application.Quit()

now=datetime.datetime.now()
cur_date=now.strftime("%Y%m%d")
#dictionary for iterating between files 
countries={9:"Austria", 3:"Europe", 6:"France", 1:"Germany", 11:"Italy", 13:"Spain", 7:"Switzerland", 2:"UK", 12:"USA", 14:"Netherlands", 10:"Belgium"}
#countries={9:"Austria"}
path='Z:\\800-Management\\830-Controlling\\833-Marketing\\Channel Controlling 2018\\'
input_file='Channel Controlling 2018 '
print('Started at: {}'.format(now.strftime("%H:%M")))
refresh_cc_sheets(path,input_file,countries)

#path='C:\\Users\\Michael\\Downloads\\'
#input_file='source_file_'
#clearing the log file
#with open(path + 'log'+ '.csv', 'w', newline='') as csvfile:
#    filewriter = csv.writer(csvfile, delimiter=',',
#                                        quotechar='|', quoting=csv.QUOTE_MINIMAL)
#    filewriter.writerow(['Log date', now])
 #   filewriter.writerow(['Filename', 'Status'])
#clearing the output file
with open(path + cur_date +'_output'+ '.csv', 'w', newline='') as csvfile:
            filewriter = csv.writer(csvfile, delimiter=',',
                                    quotechar='|', quoting=csv.QUOTE_MINIMAL)
            filewriter.writerow(['Date', 'AffiliateGroup', 'CountryId','Cost'])

for c_id, country in countries.items():


    #open the workbook
    started=datetime.datetime.now().strftime("%H:%M")
    wb = openpyxl.load_workbook(path + input_file + country + '.xlsx',read_only=True, data_only=True)

    sheets =  wb.sheetnames #list of sheet names

    #removing summary, beispiel and data pivot sheets
    i=0
    while i <=1:
        popped = sheets.pop(0)
        i+=1
    sheets.pop()
    
    #iterating between the sheets and extracting the data
    for sheet in sheets:
        #preparing lists for data that will be extracted from each sheet
        dates=[]
        costs=[]
        aff_group=[]
        country_id=[]
        data=[]
        sheet=wb[sheet]
        sheet_name=sheet[502][0].value
        # extracting only cost and date
        for row in sheet.iter_rows(min_row=3,max_row=408, min_col=1, max_col=7):
            datum = row[0].value
            cost = row[6].value
            #removing blank cells
            if datum == None:
                continue
            if cost !=None and cost != 0:
                #adding data to list
                dates.append(datum.strftime("%Y-%m-%d"))
                costs.append(cost)
                aff_group.append(sheet_name)
                country_id.append(c_id)
        #putting lists together into a list of tuples 
        data=list(zip(dates,aff_group,country_id,costs))
       
            #wiriting into the csv file
        with open(path + cur_date +'_output'+ '.csv', 'a+', newline='') as csvfile:
            filewriter = csv.writer(csvfile, delimiter=',',
                                    quotechar='|', quoting=csv.QUOTE_MINIMAL)
            for value in data:
                filewriter.writerow(value)
        #log also the sheet names for debugging        
        #with open(path + 'log'+ '.csv', 'a+', newline='') as csvfile:
        #    filewriter = csv.writer(csvfile, delimiter=',',
         #                                       quotechar='|', quoting=csv.QUOTE_MINIMAL)
         #   filewriter.writerow([sheet_name,'Done'])

    #with open(path + 'log'+ '.csv', 'a+', newline='') as csvfile:
    #    filewriter = csv.writer(csvfile, delimiter=',',
    #                                        quotechar='|', quoting=csv.QUOTE_MINIMAL)
    #    filewriter.writerow([input_file + country,'Done'])
    print('{}{} Done Started at:{} Ended at:{}'.format(input_file,country,started,datetime.datetime.now().strftime("%H:%M")))


    wb.close
now=datetime.datetime.now()
print('Ended at: {}'.format(now.strftime("%H:%M")))